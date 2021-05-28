#! /usr/bin/python3
#**********************************************************************
# Author            : Anoop Mohan
# Date created      : April 2020
#**********************************************************************
'''
	pip3 install pandas
    pip3 install openpyxl
'''

import pandas as pd
import sys 
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule


usage = "Copy the exported CSV file from Nessus to the same directory.\n Usage: python3 nessusipreport.py [csvfilename.csv]"

if len(sys.argv) <=1:
	print (usage)
	exit()

csv_export = sys.argv[1]


pd.set_option('display.max_columns', None)
template_report = "template_report.xlsx"

# Define color codes for siverities
colorcode_critical="ff0000"
colorcode_high="FFA500"
colorcode_medium="9393d3"
colorcode_low="009900"



def read_csv(csv_export):
	csvdata = pd.read_csv(csv_export)
	df_csv = pd.DataFrame(csvdata)
	df_csv.rename(columns={'Risk':'Risk Rating','Host':'IP Address', 'Name':'Vulnerability Title', 'Solution':'Recommendations'}, inplace=True)
	return (df_csv)

def open_ports(df_csv):
    portlist = df_csv[df_csv['Synopsis'] == 'It is possible to determine which TCP ports are open.']
    portlist = portlist[['IP Address','Protocol','Port']]
    portlist.reset_index(inplace=True, drop=True)
    portlist.index += 1
    return (portlist)

def write_excel(portlist):
    #work_book=load_workbook(filename='template_report.xlsx')
    work_book = load_workbook(filename=template_report)
    work_sheet=work_book['openports']
    for r in dataframe_to_rows(portlist):
            work_sheet.append(r)
    work_book.save( csv_export +'.xlsx')
  
def createsheets(unique_ip):
    work_book = load_workbook(filename=csv_export +'.xlsx')
    
    # Adding color to cells based on criticality
    #Defining the condition
    critical_background = PatternFill(bgColor=colorcode_critical)
    dxf = DifferentialStyle(fill=critical_background)
    rule = Rule(type="containsText", operator="containsText", text="Critical", dxf=dxf)
    rule.formula = ['NOT(ISERROR(SEARCH("Critical",$A1:$A1000)))']
    
    high_background = PatternFill(bgColor=colorcode_high)
    dxf = DifferentialStyle(fill=high_background)
    rule1 = Rule(type="containsText", operator="containsText", text="High", dxf=dxf)
    rule1.formula = ['NOT(ISERROR(SEARCH("High",$A1:$A1000)))']
    
    medium_background = PatternFill(bgColor=colorcode_medium)
    dxf = DifferentialStyle(fill=medium_background)
    rule2 = Rule(type="containsText", operator="containsText", text="Medium", dxf=dxf)
    rule2.formula = ['NOT(ISERROR(SEARCH("Medium",$A1:$A1000)))']
    
    low_background = PatternFill(bgColor=colorcode_low)
    dxf = DifferentialStyle(fill=low_background)
    rule3 = Rule(type="containsText", operator="containsText", text="Low", dxf=dxf)
    rule3.formula = ['NOT(ISERROR(SEARCH("Low",$A1:$A1000)))']
    
    #Applying the condition
    
    for item in unique_ip:
        work_book.create_sheet(item)
        work_book[item].conditional_formatting.add("$A1:$A1000", rule)
        work_book[item].conditional_formatting.add("$A1:$A1000", rule1)
        work_book[item].conditional_formatting.add("$A1:$A1000", rule2)
        work_book[item].conditional_formatting.add("$A1:$A1000", rule3)
    work_book.save(csv_export +'.xlsx') 
    
def remove_none(df_csv):
    df_fin = df_csv[df_csv['Risk Rating'] != "None"] 
    df_fin.reset_index(inplace=True, drop=True)
    df_fin.index += 1
    return (df_fin)
    
def ip_report(df_fin, unique_ip):   
    j=0
    df = pd.DataFrame()
    work_book = load_workbook(filename=csv_export +'.xlsx')
    for item in (df_fin['IP Address']):    
        for i in unique_ip:
            if i==item:
                df = df_fin.iloc[[j]]
     
            work_sheet=work_book[i]
            for r in dataframe_to_rows(df, index=False, header=False):
                 work_sheet.append(r)
            df.drop(df.index, inplace=True)
     
        j+=1
    work_book.save(csv_export +'.xlsx')

def xl_formatting(unique_ip):
    #Function to perform final formatting in the Xl sheet
    work_boo = load_workbook(filename=csv_export +'.xlsx')
    for ip in unique_ip:
        work_sh = work_boo[ip]
        work_sh.insert_rows(0)
        work_sh.cell(row=1, column=1).value = "Risk Rating"
        work_sh.cell(row=1, column=2).value = "IP Address"
        work_sh.cell(row=1, column=3).value = "Protocol"
        work_sh.cell(row=1, column=4).value = "Port"
        work_sh.cell(row=1, column=5).value = "Vulnerability Title"
        work_sh.cell(row=1, column=6).value = "Score"
        work_sh.cell(row=1, column=7).value = "Description"
        work_sh.cell(row=1, column=8).value = "Recommendations"
        work_sh.cell(row=1, column=9).value = "Plugin Output"
        work_sh.column_dimensions['A'].width='8.57'
        work_sh.column_dimensions['B'].width='9.71'
        work_sh.column_dimensions['C'].width='7.71'
        work_sh.column_dimensions['D'].width='6.57'
        work_sh.column_dimensions['E'].width='17.29'
        work_sh.column_dimensions['F'].width='5.14'
        work_sh.column_dimensions['G'].width='63.71'
        work_sh.column_dimensions['H'].width='18'
        work_sh.column_dimensions['I'].width='80.86'
        i = range(1,10)
        for x in i:
            work_sh.cell(row=1, column=x).border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            work_sh.cell(row=1, column=x).font = Font(b=True, color="ffffff")
            work_sh.cell(row=1, column=x).fill = PatternFill("solid", fgColor="2F75B5")
    
    
        for eachrow in work_sh.iter_rows():
            for cell in eachrow:
                wrap_alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')
                cell.alignment = wrap_alignment
                cell.border =  Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        work_boo.save(csv_export +'.xlsx')
	 


df_csv = read_csv(csv_export)
df_csv = df_csv.drop_duplicates()
df_fin = remove_none(df_csv)
df_fin = df_fin[['Risk Rating', 'IP Address', 'Protocol', 'Port', 'Vulnerability Title', 'CVSS', 'Description', 'Recommendations','Plugin Output']]
df_fin = df_fin.drop_duplicates()
df_fin = df_fin.sort_values(by = 'CVSS', ascending=False)
df_fin = df_fin.reset_index(drop=True)
unique_ip = (df_fin['IP Address'].unique())

portlist = open_ports(df_csv)
write_excel(portlist)
createsheets(unique_ip)

ip_report(df_fin,unique_ip)
xl_formatting(unique_ip)

  


   
   

